import sys
import openpyxl
import xlsxwriter
import regex
import pandas as pd
from window import *


global select_file, output_txt
global select_file_2, output_xlsx

def geo_xlsx_to_txt():
    global select_file, output_txt  
    book = openpyxl.open(select_file, read_only=True)
    sheet = book.active

    output_file = open(output_txt,'w')

    for row in range (3,sheet.max_row+1):
        coord_line = str(
            sheet[row][0].value)+'; '+str(
                sheet[row][1].value)+' '+str(
                    sheet[row][2].value)+' '+str(
                        sheet[row][3].value)+'; '+str(
                            sheet[row][4].value)+' '+str(
                                sheet[row][5].value)+' '+str(
                                    sheet[row][6].value)+';\n'

        output_file.write(coord_line)
    output_file.close()

    open_file = open(output_txt).read()
    regular_ka = r'\1.\2'
    fine_text = regex.sub(r'(\d{0})+[.,]+(\d{0})',regular_ka ,open_file)

    wr = open(output_txt,'w')
    wr.write(fine_text)
    wr.close()
    
def xy_xlsx_to_txt():
    global select_file, output_txt
    book = openpyxl.open(select_file, read_only=True)
    sheet = book.active

    output_file = open(output_txt,'w')

    for row in range (2,sheet.max_row+1):
        coord_line = str(
            sheet[row][0].value)+'; '+str(
                sheet[row][1].value)+'; '+str(
                    sheet[row][2].value)+';\n'

        output_file.write(coord_line)
    output_file.close()

    open_file = open(output_txt).read()
    regular_ka = r'\1.\2'
    fine_text = regex.sub(r'(\d{0})+[.,]+(\d{0})',regular_ka ,open_file)

    wr = open(output_txt,'w')
    wr.write(fine_text)
    wr.close()

def xy_txt_to_xlsx():
    global select_file_2, output_xlsx
    open_file = open(select_file_2).read()
    cleaner = r'\1.\2'
    fine_text = regex.sub(r'(\d{0})+[.,]+(\d{0})',cleaner ,open_file)

    wr = open(select_file_2,'w')
    wr.write(fine_text)
    wr.close()

    data = pd.read_csv(select_file_2, encoding='cp1251', header=None, sep=r"\s+", names=None, )
    data=data.replace({';': ''}, regex=True)

    workbook = xlsxwriter.Workbook(output_xlsx)
    worksheet = workbook.add_worksheet('Coordinates')

    worksheet.set_column('B:C', 16.22)
    worksheet.set_column('D:E', 24.33)

    write_format_hat = workbook.add_format({
            'font': 'Times New Roman',
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': 'white'})

    write_format = workbook.add_format({
            'bold': 0,
            'border': 1,
            'num_format': "0.000",
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': 'white'})


    worksheet.write('A1', '№№', write_format_hat)
    worksheet.write('B1', 'X', write_format_hat)
    worksheet.write('C1', 'Y', write_format_hat)

    a = 0
    data_shape_2 = data.shape[0] - 1
    bcolmn = 'B'
    b = 0
    ccolmn = 'C'
    c = 0

    while a < data_shape_2:
        vertical = ('A' + str(a + 2))
        bb = (bcolmn + str(b + 2))
        cc = (ccolmn + str(c + 2))

        iat_data_2 = float(data.iat[a, 2])
        iat_data_1 = float(data.iat[a, 1])

        worksheet.write(bb, iat_data_2, write_format)
        worksheet.write(cc, iat_data_1, write_format) 

        a += 1
        b += 1 
        c += 1

        worksheet.write(vertical, a, write_format_hat)

    workbook.close()

def geo_txt_to_xlsx():
    global select_file_2, output_xlsx
    open_file = open(select_file_2).read()
    cleaner = r'\1.\2'
    fine_text = regex.sub(r'(\d{0})+[.,]+(\d{0})',cleaner ,open_file)

    wr = open(select_file_2,'w')
    wr.write(fine_text)
    wr.close()

    data = pd.read_csv(select_file_2, encoding='cp1251', header=None, sep=r"\s+", names=None, )
    data=data.replace({';': ''}, regex=True)

    workbook = xlsxwriter.Workbook(output_xlsx)
    worksheet = workbook.add_worksheet('Coordinates')

    # worksheet.set_column('D:D', 12)
    # worksheet.set_column('G:G', 12)

    merge_format = workbook.add_format({
            'bold': 0,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': 'white'})

    write_format = workbook.add_format({
            'bold': 0,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': 'white'})

    write_format_colum = workbook.add_format({
            'bold': 0,
            'border': 1,
            'num_format': "00",
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': 'white'})

    write_format1 = workbook.add_format({
            'bold': 0,
            'border': 1,
            'num_format': "00.00000",
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': 'white'})

    worksheet.merge_range('A1:A2', '№№', merge_format)
    worksheet.merge_range('B1:D1', 'North Latitude', merge_format)
    worksheet.merge_range('E1:G1', 'East Longitude', merge_format)
    worksheet.write('B2', 'Degrees', write_format)
    worksheet.write('C2', 'Minutes', write_format)
    worksheet.write('D2', 'Seconds', write_format)
    worksheet.write('E2', 'Degrees', write_format)
    worksheet.write('F2', 'Minutes', write_format)
    worksheet.write('G2', 'Seconds', write_format)
    worksheet.write('A1', '№№', write_format1)

    a = 0
    data_shape = data.shape[0]
    bcolmn = 'B'
    b = 0
    ccolmn = 'C'
    c = 0
    dcolmn = 'D'
    d = 0
    ecolmn = 'E'
    e = 0
    fcolmn = 'F'
    f = 0
    gcolmn = 'G'
    g = 0
    yy = -1

    while a < data_shape:
        a += 1
        b += 1
        c += 1
        d += 1
        e += 1
        f += 1
        g += 1
        yy += 1
        vertical_2 = ('A' + str(a + 2))
        bb = (bcolmn + str(b + 2))
        cc = (ccolmn + str(c + 2))
        dd = (dcolmn + str(d + 2))
        ee = (ecolmn + str(e + 2))
        ff = (fcolmn + str(f + 2))
        gg = (gcolmn + str(g + 2))

        worksheet.write(vertical_2, a, write_format)
        
        float_iat_2 = float(data.iat[yy, 2])
        float_iat_3 = float(data.iat[yy, 3])
        float_iat_5 = float(data.iat[yy, 5])
        float_iat_6 = float(data.iat[yy, 6])

        worksheet.write(bb, data.iat[yy, 1], write_format)
        worksheet.write(cc, float_iat_2, write_format_colum)
        worksheet.write(dd, float_iat_3, write_format1)
        worksheet.write(ee, data.iat[yy, 4], write_format)
        worksheet.write(ff, float_iat_5, write_format_colum)
        worksheet.write(gg, float_iat_6, write_format1)

    workbook.close()

class gce(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_gce()
        self.ui.setupUi(self)
        
        self.ui.pushButton_3.clicked.connect(self.file_selection)
        self.ui.pushButton.clicked.connect(self.save_file_geo)
        self.ui.pushButton_2.clicked.connect(self.save_file_xy)
        self.ui.pushButton_4.clicked.connect(self.file_selection_2)
        self.ui.pushButton_10.clicked.connect(self.save_file_xlsx_xy)
        self.ui.pushButton_9.clicked.connect(self.save_file_xlsx_geo)
    
    def file_selection(self):
        global select_file
        self.label_status_calm()
        select_file = QFileDialog.getOpenFileName(self, "File selection", "Your file", "text (*.xlsx)")[0]
        if not select_file:
            QtWidgets.QMessageBox.about(self, 'Warning', 'Need to select a .xlsx file!')
        else:
            self.path_label()
            self.ui.pushButton.setDisabled(False)
            self.ui.pushButton_2.setDisabled(False)
        return select_file
    
    def file_selection_2(self):
        global select_file_2
        self.label_status_calm_2()
        select_file_2 = QFileDialog.getOpenFileName(self, "File selection", "Your file", "text (*.txt)")[0]
        if not select_file_2:
            QtWidgets.QMessageBox.about(self, 'Warning', 'Need to select a .txt file!')
        else:
            self.path_label_2()
            self.ui.pushButton_9.setDisabled(False)
            self.ui.pushButton_10.setDisabled(False)
        return select_file_2

    def save_file_xlsx_geo(self):
        global output_xlsx
        output_xlsx = False
        output_xlsx = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "Your file", "*.xlsx")[0]
        if not output_xlsx:
            QtWidgets.QMessageBox.about(self, 'Warning', 'Need to select the name and path of the file!')
        else:
            geo_txt_to_xlsx()
            self.finish_label_2()
            self.ui.pushButton_9.setDisabled(True)
            self.ui.pushButton_10.setDisabled(True)

    def save_file_xlsx_xy(self):
        global output_xlsx
        output_xlsx = False
        output_xlsx = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "Your file", "*.xlsx")[0]
        if not output_xlsx:
            QtWidgets.QMessageBox.about(self, 'Warning', 'Need to select the name and path of the file!')
        else:
            xy_txt_to_xlsx()
            self.finish_label_2()
            self.ui.pushButton_9.setDisabled(True)
            self.ui.pushButton_10.setDisabled(True)
    
    def save_file_geo(self):
        global output_txt
        output_txt = False
        output_txt = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "Your file", "*.txt")[0]
        if not output_txt:
            QtWidgets.QMessageBox.about(self, 'Warning', 'Need to select the name and path of the file!')
        else:
            geo_xlsx_to_txt()
            self.finish_label()
            self.ui.pushButton.setDisabled(True)
            self.ui.pushButton_2.setDisabled(True)

    def save_file_xy(self):
        global output_txt
        output_txt = False
        output_txt = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "Your file", "*.txt")[0]
        if not output_txt:
            QtWidgets.QMessageBox.about(self, 'Warning', 'Need to select the name and path of the file!')
        else:
            xy_xlsx_to_txt()
            self.finish_label()
            self.ui.pushButton.setDisabled(True)
            self.ui.pushButton_2.setDisabled(True)

    def path_label(self):
        self.ui.label_2.setText(select_file)
    
    def path_label_2(self):
        self.ui.label_5.setText(select_file_2)

    def finish_label(self):
        self.ui.label_2.setText('Finish')

    def finish_label_2(self):
        self.ui.label_5.setText('Finish')

    def label_status_calm(self):
        self.ui.label_2.setText('Status')
    
    def label_status_calm_2(self):
        self.ui.label_5.setText('Status')


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    myapp = gce()
    myapp.show()
    sys.exit(app.exec_())